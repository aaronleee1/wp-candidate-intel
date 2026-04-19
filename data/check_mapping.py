import json
from pathlib import Path
from openpyxl import load_workbook

z2c = json.loads(Path('data/zip-county.json').read_text())
bsa = json.loads(Path('data/bsa-councils.json').read_text())

bsa_zips = sorted({str(x.get('zip', '')).zfill(5) for x in bsa if x.get('zip')})
bsa_hit = sum(1 for z in bsa_zips if z in z2c)

wb = load_workbook('data/jrotc.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

header_row_idx = 0
for i, row in enumerate(rows[:12]):
    cells = [str(c or '').strip().lower() for c in row]
    if any('zip' in c for c in cells) and any('school' in c for c in cells):
        header_row_idx = i
        break

headers = [str(c or '').strip().lower() for c in rows[header_row_idx]]
zip_col = next((i for i, h in enumerate(headers) if 'zip' in h), -1)

jz = set()
for r in rows[header_row_idx + 1:]:
    v = '' if zip_col < 0 else str(r[zip_col] or '')
    digits = ''.join(ch for ch in v if ch.isdigit())[:5]
    if digits:
        jz.add(digits.zfill(5))

j_hit = sum(1 for z in jz if z in z2c)

print('zip-county entries:', len(z2c))
print('bsa zips:', len(bsa_zips), 'mapped:', bsa_hit)
print('jrotc zips:', len(jz), 'mapped:', j_hit)
