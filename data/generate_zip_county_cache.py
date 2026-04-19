import json
import ssl
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook

ctx = ssl._create_unverified_context()


def get_json(url: str, timeout: int = 12):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.loads(resp.read().decode("utf-8", "ignore"))


zips = set()

# JROTC zips
wb = load_workbook("data/jrotc.xlsx", read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header_idx = None
zip_col = None
for i, row in enumerate(rows[:10]):
    cells = [str(c).lower().strip() if c is not None else "" for c in row]
    if any("zip" in c for c in cells) and any("school" in c for c in cells):
        header_idx = i
        zip_col = next((j for j, h in enumerate(cells) if "zip" in h), None)
        break
if header_idx is not None and zip_col is not None:
    for row in rows[header_idx + 1 :]:
        if zip_col >= len(row):
            continue
        raw = row[zip_col]
        digits = "".join(ch for ch in str(raw) if ch.isdigit())[:5]
        if len(digits) == 5:
            zips.add(digits)

# BSA zips
for c in json.loads(Path("data/bsa-councils.json").read_text()):
    digits = "".join(ch for ch in str(c.get("zip", "")) if ch.isdigit())[:5]
    if len(digits) == 5:
        zips.add(digits)

county_map = {}


def resolve_zip(z: str):
    try:
        zp = get_json(f"https://api.zippopotam.us/us/{z}")
        place = (zp.get("places") or [{}])[0]
        lat = place.get("latitude")
        lng = place.get("longitude")
        if lat is None or lng is None:
            return z, ""
        area = get_json(f"https://geo.fcc.gov/api/census/area?lat={lat}&lon={lng}&format=json")
        fips = (area.get("results") or [{}])[0].get("county_fips") or ""
        return z, str(fips).zfill(5) if fips else ""
    except Exception:
        return z, ""


with ThreadPoolExecutor(max_workers=16) as ex:
    futures = [ex.submit(resolve_zip, z) for z in sorted(zips)]
    for fut in as_completed(futures):
        z, fips = fut.result()
        if fips:
            county_map[z] = fips

Path("data/zip-county.json").write_text(json.dumps(county_map, indent=2, sort_keys=True))
print(f"zips: {len(zips)}")
print(f"mapped: {len(county_map)}")
